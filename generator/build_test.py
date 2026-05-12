"""
generator.build_test — Generate a pytest suite from a recording.jsonl artifact.

Produces in <out_dir>/:
  test_<name>.py    — parametrized pytest file importing Page Objects
    conftest.py       — session fixtures (playwright_page, java_driver, healer)
  data.xlsx         — Excel template with placeholder columns pre-populated
  README.txt        — columns, run command, originating English steps
"""
from __future__ import annotations

import json
import builtins
import keyword
import pprint
import re
import sys
from pathlib import Path
from typing import Any

import config
from qcs_repo import store as repo_store

# ── Helpers ───────────────────────────────────────────────────────────────────

def _snake(name: str) -> str:
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "_", name)
    s = re.sub(r"[^a-z0-9_]", "_", s.lower())
    return re.sub(r"_+", "_", s).strip("_")


def _class_name(form_id: str) -> str:
    slug = form_id.removeprefix("java_").removeprefix("html_")
    return "".join(w.capitalize() for w in slug.split("_")) + "Page"


def _load_recording(run_dir: Path) -> list[dict]:
    jsonl = run_dir / "recording.jsonl"
    if not jsonl.exists():
        raise FileNotFoundError(f"recording.jsonl not found in {run_dir}")
    rows = []
    for line in jsonl.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            rows.append(json.loads(line))
    return rows


def _is_raw_java_ref(ref: str) -> bool:
    return re.fullmatch(r"e\d+", ref or "") is not None


def _form_var(form_id: str) -> str:
    slug = form_id.removeprefix("java_").removeprefix("html_")
    if not slug:
        return "form"
    if "order" in slug or "quote" in slug:
        return "order_form"
    return _snake(slug) + "_form"


def _form_var_from_label(label: str, fallback_form_id: str = "") -> str:
    text = re.sub(r"\b(form|screen)\b", "", label or "", flags=re.I)
    text = re.sub(r"\bfor\b", "", text, flags=re.I)
    variable = _snake(text) or _form_var(fallback_form_id)
    if not variable:
        return "oracle_form"
    return variable


def _form_title_from_repo(form_id: str) -> str:
    form = repo_store.load_form(form_id) or {}
    title = str(form.get("title") or form.get("name") or "").strip()
    if title and title != form_id:
        return title
    return form_id.removeprefix("java_").removeprefix("html_")


def _form_label_from_repo(form_id: str, fallback_ref: str = "") -> str:
    form = repo_store.load_form(form_id) or {}
    raw_title = str(form.get("title") or form.get("name") or "").strip()
    derived_from_id = not raw_title or raw_title == form_id
    title = raw_title if not derived_from_id else form_id.removeprefix("java_").removeprefix("html_")
    label = _clean_form_label(title, form_id)
    if label.lower() in {"yes", "no", "ok", "cancel"}:
        return f"{label} Dialog"
    if derived_from_id and label.lower() in {"find", "open", "save", "search"}:
        return f"{label} Window"
    return label


def _unique_form_label(base_label: str, used_labels: set[str]) -> str:
    label = base_label or "Oracle Form"
    if label not in used_labels:
        used_labels.add(label)
        return label
    index = 2
    while f"{label} {index}" in used_labels:
        index += 1
    label = f"{label} {index}"
    used_labels.add(label)
    return label


def _unique_var(base_var: str, used_vars: set[str]) -> str:
    variable = base_var or "oracle_form"
    if not re.match(r"[a-zA-Z_]", variable):
        variable = f"form_{variable}"
    if keyword.iskeyword(variable) or variable in dir(builtins):
        variable = f"{variable}_form"
    if variable not in used_vars:
        used_vars.add(variable)
        return variable
    index = 2
    while f"{variable}_{index}" in used_vars:
        index += 1
    variable = f"{variable}_{index}"
    used_vars.add(variable)
    return variable


def _resolve_recorded_element(form_id: str, ref: str) -> tuple[str, dict] | None:
    if not ref:
        return None
    if form_id:
        resolved = repo_store.resolve_element_ref(ref, form_id=form_id)
        if resolved:
            return resolved
    try:
        return repo_store.resolve_element_ref(ref)
    except RuntimeError:
        return None


def _control_owner_form_id(form_id: str, ref: str) -> str:
    resolved = _resolve_recorded_element(form_id, ref)
    if resolved:
        return resolved[0]
    return form_id


def _clean_label(raw: str, fallback_ref: str = "") -> str:
    label = raw or fallback_ref or "element"
    label = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", label)
    label = re.sub(r"[_\-]+", " ", label)
    label = re.sub(r"list\s+of\s+values", "", label, flags=re.I)
    label = re.sub(r"\balt\s+[a-z0-9]\b", "", label, flags=re.I)
    label = re.sub(r"\bmnemonic\s+[a-z0-9]\b", "", label, flags=re.I)
    label = re.sub(r"\s+", " ", label).strip()
    if not label:
        label = fallback_ref or "element"
    words = []
    for word in label.split(" "):
        if word.isupper() or "#" in word:
            words.append(word)
        else:
            words.append(word[:1].upper() + word[1:])
    return " ".join(words)


def _clean_form_label(raw: str, fallback_ref: str = "") -> str:
    label = raw or fallback_ref or "Oracle Form"
    label = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", label)
    label = re.sub(r"[_\-]+", " ", label)
    label = re.sub(r"\balt\s+[a-z0-9]\b", "", label, flags=re.I)
    label = re.sub(r"\bmnemonic\s+[a-z0-9]\b", "", label, flags=re.I)
    label = re.sub(r"\s+", " ", label).strip()
    if not label:
        label = fallback_ref or "Oracle Form"
    return " ".join(
        word if word.isupper() or "#" in word else word[:1].upper() + word[1:]
        for word in label.split(" ")
    )


def _element_label(form_id: str, ref: str, *, method: str | None = None) -> str:
    resolved = _resolve_recorded_element(form_id, ref)
    element = resolved[1] if resolved else {}
    raw = str(
        element.get("description")
        or element.get("name")
        or element.get("friendly_name")
        or ref
    )
    label = _clean_label(raw, ref)
    if method == "toolbar" and re.fullmatch(r"toolbar\d+", label.replace(" ", ""), flags=re.I):
        return "Toolbar"
    return label


def _control_method(form_id: str, ref: str) -> str:
    resolved = _resolve_recorded_element(form_id, ref)
    element = resolved[1] if resolved else {}
    role = str(element.get("role") or "").lower()
    if "field" in role or "text" in role:
        return "textbox"
    if "button" in role:
        return "button"
    if "menu" in role:
        return "menu"
    if "toolbar" in role:
        return "toolbar"
    return "element"


def _control_expr(form_var: str, label: str, method: str) -> str:
    return f"{form_var}.{method}({label!r})"


def _control_kind(method: str) -> str:
    return "field" if method == "textbox" else method


def _register_object(
    objects: dict[str, dict[str, Any]],
    *,
    form_id: str,
    ref: str,
    method: str,
) -> str:
    label = _element_label(form_id, ref, method=method)
    entry = {"kind": _control_kind(method), "form_id": form_id, "ref": ref}
    existing = objects.get(label)
    if existing == entry:
        return label
    if existing is None:
        objects[label] = entry
        return label

    readable_context = _clean_label(form_id.removeprefix("java_").removeprefix("html_"), form_id)
    contextual_label = f"{label} In {readable_context}"
    suffix = 2
    candidate = contextual_label
    while candidate in objects and objects[candidate] != entry:
        suffix += 1
        candidate = f"{contextual_label} {suffix}"
    objects[candidate] = entry
    return candidate


def _form_label(search_text: str, form_name: str) -> str:
    text = re.sub(r"^\s*\d+\.\s*", "", search_text or "").strip()
    match = re.search(r"\bopen\s+(.+)$", text, flags=re.I)
    if match:
        return _clean_label(match.group(1), form_name or "Oracle Form")
    return _clean_label(text or form_name or "Oracle Form")


# ── Generate test file ────────────────────────────────────────────────────────

def generate_test(run_dir: Path, out_dir: Path, test_name: str) -> None:
    rows = _load_recording(run_dir)

    # Collect info
    placeholders: list[dict]           = []   # {column_name, sample_value}
    assertions:   list[dict]           = []
    form_ids:     set[str]             = set()
    page_form_ids: set[str]            = set()
    flow_ids:     set[str]             = set()
    steps:        list[tuple[str,dict]] = []  # (op, row)

    for row in rows:
        op = row.get("op", "")
        if op == "data_placeholder":
            placeholders.append(row)
        elif op == "assertion":
            steps.append((op, row))
        elif op == "repo_invoke_flow":
            flow_ids.add(row.get("flow_id", ""))
        elif op in ("java_send_text", "java_click", "java_press_key", "java_form_close"):
            tgt = row.get("target", {})
            if isinstance(tgt, dict):
                fid = tgt.get("form_id", "")
                friendly_name = tgt.get("friendly_name", "")
                if fid:
                    form_ids.add(fid)
                    if (
                        friendly_name
                        and not _is_raw_java_ref(friendly_name)
                        and repo_store.find_element(fid, friendly_name)
                    ):
                        page_form_ids.add(fid)
            steps.append((op, row))
        elif op == "java_form_launch":
            fid = row.get("form_id", "")
            if fid:
                form_ids.add(fid)
            steps.append((op, row))
        elif op in ("oracle_form_open", "ebs_login"):
            steps.append((op, row))

    has_placeholders     = bool(placeholders)
    has_java             = any(row.get("surface") == "java" for row in rows)
    has_html             = any(
        row.get("surface") == "html" or row.get("op") == "ebs_login"
        for row in rows
    )
    has_java_form_launch = any(op == "java_form_launch" for op, _ in steps)
    has_assertions       = any(op == "assertion" for op, _ in steps)

    # ── Imports ───────────────────────────────────────────────────────────────
    lines = [
        "# AUTO-GENERATED — do not edit the step section by hand.",
        "# Update the repository/generator source and rerun `qcs gen` for lasting changes.",
        "from __future__ import annotations",
    ]

    if has_placeholders:
        lines.append("import pytest")
        lines.append("from qcs_replay.data import load_excel_rows")

    if has_assertions and has_java:
        lines.append("from qcs_replay.assertions import assert_java_text, assert_java_state")
    if has_assertions and has_html:
        lines.append("from qcs_replay.assertions import assert_pw_text, assert_pw_value")
    if has_java_form_launch:
        lines.append("from qcs_replay.script import OracleReplay")
    elif has_html:
        lines.append("import os")

    # Flow imports
    for fid in sorted(flow_ids):
        fn = _snake(fid)
        lines.append(f"from flows_gen.{fid} import {fn}")

    lines += ["", ""]

    # ── Parametrize decorator ─────────────────────────────────────────────────
    if has_placeholders:
        lines += [
            "@pytest.mark.parametrize(",
            '    "data_row",',
            '    load_excel_rows(str(__file__).replace(".py", "_data.xlsx")),',
            "    ids=lambda r: str(r.get('_id', '')),",
            ")",
    ]

    # ── Test function ─────────────────────────────────────────────────────────
    fn_args = "data_row, " if has_placeholders else ""
    if has_java_form_launch:
        fn_args += "oracle: OracleReplay, "
    elif has_html:
        fn_args += "playwright_page, "
    if has_java and not has_java_form_launch:
        fn_args += "java_driver, "
    if has_java_form_launch:
        fn_args = fn_args.rstrip(", ")
    else:
        fn_args += "healer"

    lines += [
        f"def test_{_snake(test_name)}({fn_args}):",
        f'    """Generated from: {run_dir.name}"""',
    ]
    lines.append("")

    # Emit readable replay steps.
    active_java_form_id = ""
    object_repository: dict[str, dict[str, Any]] = {}
    form_labels_by_id: dict[str, str] = {}
    used_form_labels: set[str] = set()
    pending_form_open: dict[str, Any] = {}

    def ensure_form_handle(
        form_id: str,
        *,
        fallback_ref: str = "",
        open_form: bool = False,
        url: str = "",
        expected_name: str = "",
        preferred_label: str = "",
    ) -> tuple[str, dict[str, Any], str]:
        if not form_id:
            form_id = active_java_form_id
        if not form_id:
            raise RuntimeError("Cannot generate Java step without a form_id")

        if form_id not in form_labels_by_id:
            base_label = preferred_label or _form_label_from_repo(form_id, fallback_ref)
            form_labels_by_id[form_id] = _unique_form_label(base_label, used_form_labels)
        label = form_labels_by_id[form_id]

        objects = object_repository.setdefault(
            label,
            {
                "__form__": {
                    "form_id": form_id,
                    "expected_name": expected_name or _form_title_from_repo(form_id),
                }
            },
        )
        objects.setdefault("__form__", {"form_id": form_id})
        objects["__form__"].setdefault("form_id", form_id)
        if expected_name:
            objects["__form__"]["expected_name"] = expected_name

        if open_form:
            lines.append(f"    oracle.open_form(url={url!r})")
        return f"oracle.form({label!r})", objects, label

    i = 0
    while i < len(steps):
        op, row = steps[i]
        tgt = row.get("target", {})
        fn = tgt.get("friendly_name", "") if isinstance(tgt, dict) else str(tgt)
        fid = tgt.get("form_id", "") if isinstance(tgt, dict) else ""

        if op == "java_click" and i + 2 < len(steps):
            next_op, next_row = steps[i + 1]
            final_op, final_row = steps[i + 2]
            final_tgt = final_row.get("target", {})
            final_ref = final_tgt.get("friendly_name", "") if isinstance(final_tgt, dict) else ""
            final_fid = final_tgt.get("form_id", "") if isinstance(final_tgt, dict) else ""
            if (
                next_op == "java_press_key"
                and str(next_row.get("key", "")).upper() == "CTRL+A"
                and final_op == "java_send_text"
                and final_ref == fn
            ):
                owner_form_id = _control_owner_form_id(final_fid or fid or active_java_form_id, final_ref)
                active_java_form_id = owner_form_id
                form_expr, active_objects, _ = ensure_form_handle(
                    owner_form_id,
                    fallback_ref=final_ref,
                )
                label = _register_object(
                    active_objects,
                    form_id=owner_form_id,
                    ref=final_ref,
                    method="textbox",
                )
                control = _control_expr(form_expr, label, "textbox")
                text = final_row.get("text", "")
                ph = next((p for p in placeholders if p.get("sample_value") == text), None)
                lines.append(f"    {control}.clear()")
                if ph:
                    lines.append(f'    {control}.set(data_row[{ph["column_name"]!r}])')
                else:
                    lines.append(f"    {control}.set({text!r})")
                i += 3
                continue

        if op == "java_send_text":
            text = row.get("text", "")
            owner_form_id = _control_owner_form_id(fid or active_java_form_id, fn)
            if owner_form_id:
                active_java_form_id = owner_form_id
            ph = next((p for p in placeholders if p.get("sample_value") == text), None)
            form_expr, active_objects, _ = ensure_form_handle(
                owner_form_id or active_java_form_id,
                fallback_ref=fn,
            )
            label = _register_object(
                active_objects,
                form_id=owner_form_id or active_java_form_id,
                ref=fn,
                method="textbox",
            )
            control = _control_expr(form_expr, label, "textbox")
            if ph:
                lines.append(f'    {control}.set(data_row[{ph["column_name"]!r}])')
            else:
                lines.append(f"    {control}.set({text!r})")

        elif op == "java_click":
            owner_form_id = _control_owner_form_id(fid or active_java_form_id, fn)
            if owner_form_id:
                active_java_form_id = owner_form_id
            form_expr, active_objects, _ = ensure_form_handle(
                owner_form_id or active_java_form_id,
                fallback_ref=fn,
            )
            method = _control_method(owner_form_id or active_java_form_id, fn)
            label = _register_object(
                active_objects,
                form_id=owner_form_id or active_java_form_id,
                ref=fn,
                method=method,
            )
            control = _control_expr(form_expr, label, method)
            lines.append(f"    {control}.click()")

        elif op == "java_form_close":
            form_expr, _, _ = ensure_form_handle(active_java_form_id)
            lines.append(f"    {form_expr}.close()")

        elif op == "java_press_key":
            key = row.get("key", "")
            form_expr, _, _ = ensure_form_handle(active_java_form_id)
            lines.append(f"    {form_expr}.press_key({key!r})")

        elif op == "assertion":
            target = row.get("target", "element")
            expected_text = row.get("expected_text")
            expected_state = row.get("expected_state")
            owner_form_id = _control_owner_form_id(active_java_form_id, str(target))
            method = _control_method(owner_form_id or active_java_form_id, str(target))
            form_expr, active_objects, _ = ensure_form_handle(
                owner_form_id or active_java_form_id,
                fallback_ref=str(target),
            )
            label = _register_object(
                active_objects,
                form_id=owner_form_id or active_java_form_id,
                ref=str(target),
                method=method,
            )
            control = _control_expr(form_expr, label, method)
            if expected_text:
                lines.append(f"    assert_java_text({control}, {expected_text!r})")
            if expected_state:
                lines.append(f"    assert_java_state({control}, {expected_state!r})")

        elif op == "ebs_login":
            url = row.get("url", "")
            user_env = row.get("user_env", "EBS_USER")
            pass_env = row.get("password_env", "EBS_PASSWORD")
            lines.append(
                f"    oracle.login(url={url!r}, user_env={user_env!r}, password_env={pass_env!r})"
            )

        elif op == "java_form_launch":
            url = row.get("url", "")
            form_name = row.get("form_name", "")
            active_java_form_id = row.get("form_id", "") or active_java_form_id
            lines.append(f"    oracle.open_form(url={url!r})")
            ensure_form_handle(
                active_java_form_id,
                expected_name=form_name,
            )

        elif op == "oracle_form_open":
            pending_form_open = row
            url = row.get("url", "")
            if url and not has_java_form_launch:
                lines.append(f"    playwright_page.goto({url!r})")

        i += 1

    lines += ["", ""]

    # ── Write test file ───────────────────────────────────────────────────────
    out_dir.mkdir(parents=True, exist_ok=True)
    test_file = out_dir / f"test_{_snake(test_name)}.py"
    test_file.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Test file: {test_file}")

    if object_repository:
        _write_object_repository(out_dir, object_repository)

    # ── Write conftest ────────────────────────────────────────────────────────
    _write_conftest(
        out_dir,
        has_java=has_java,
        has_html=has_html,
        has_java_form_launch=has_java_form_launch,
        run_dir=run_dir,
    )

    # ── Write Excel template ──────────────────────────────────────────────────
    if has_placeholders:
        _write_excel(out_dir, test_name, placeholders)

    # ── Write README ──────────────────────────────────────────────────────────
    _write_readme(out_dir, test_name, placeholders, run_dir)

    print(f"  Done → {out_dir}")


def _write_object_repository(out_dir: Path, object_repository: dict[str, dict[str, Any]]) -> None:
    lines = [
        "# AUTO-GENERATED object repository for readable replay scripts.",
        "from __future__ import annotations",
        "",
        "OBJECT_REPOSITORY = ",
    ]
    text = pprint.pformat(object_repository, width=100, sort_dicts=False)
    (out_dir / "objects.py").write_text("\n".join(lines) + text + "\n", encoding="utf-8")
    print(f"  Object repository: {out_dir / 'objects.py'}")


# ── Conftest ──────────────────────────────────────────────────────────────────

def _write_conftest(
    out_dir: Path,
    has_java: bool,
    has_html: bool,
    has_java_form_launch: bool,
    run_dir: Path,
) -> None:
    lines = [
        "# AUTO-GENERATED conftest.py",
        "from __future__ import annotations",
        "import os",
        "from pathlib import Path",
        "import pytest",
        "import config",
        "from qcs_replay.healing import HealingEngine",
        "",
    ]

    if has_html or has_java_form_launch:
        lines += [
            "@pytest.fixture(scope='session')",
            "def playwright_page():",
            "    from playwright.sync_api import sync_playwright",
            "    with sync_playwright() as pw:",
            "        browser = pw.chromium.launch(headless=False)",
            "        ctx = browser.new_context(accept_downloads=True)",
            "        page = ctx.new_page()",
            "        yield page",
            "        browser.close()",
            "",
        ]

    if has_java_form_launch:
        lines += [
            "@pytest.fixture",
            "def oracle(playwright_page, healer):",
            "    from qcs_replay.script import OracleReplay",
            "    from objects import OBJECT_REPOSITORY",
            "    return OracleReplay(playwright_page, healer=healer, object_repository=OBJECT_REPOSITORY)",
            "",
        ]

    if has_java and not has_java_form_launch:
        lines += [
            "@pytest.fixture(scope='session')",
            "def java_driver():",
            "    from qcs_replay.java_agent import attach_java_agent",
            "    driver = attach_java_agent()",
            "    yield driver",
            "",
        ]

    lines += [
        "@pytest.fixture(autouse=True)",
        "def healer(request, tmp_path):",
        f"    patch_path = Path({str(run_dir)!r}) / 'repo_patch.yaml'",
        "    engine = HealingEngine(patch_path=patch_path)",
        "    yield engine",
        "    # Attach healing summary to pytest report",
        "    for ev in engine.events:",
        "        request.node.user_properties.append((",
        "            f'heal_{ev.tier}', f'{ev.step_intent}→success={ev.success}',",
        "        ))",
        "",
    ]

    (out_dir / "conftest.py").write_text("\n".join(lines), encoding="utf-8")
    print(f"  conftest.py: {out_dir / 'conftest.py'}")


# ── Excel template ────────────────────────────────────────────────────────────

def _write_excel(out_dir: Path, test_name: str, placeholders: list[dict]) -> None:
    import openpyxl  # noqa: PLC0415
    from openpyxl.styles import Font, PatternFill  # noqa: PLC0415

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl did not create an active worksheet")
    ws.title = "TestData"

    # Headers
    cols = [p["column_name"] for p in placeholders]
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    # Pre-populate sample row
    for i, p in enumerate(placeholders, 1):
        ws.cell(row=2, column=i, value=p.get("sample_value", ""))

    xl_path = out_dir / f"test_{_snake(test_name)}_data.xlsx"
    wb.save(xl_path)
    print(f"  Excel template: {xl_path}")


# ── README ────────────────────────────────────────────────────────────────────

def _write_readme(
    out_dir: Path,
    test_name: str,
    placeholders: list[dict],
    run_dir: Path,
) -> None:
    lines = [
        f"Test: {test_name}",
        f"Generated from recording: {run_dir.name}",
        "",
        "Run command:",
        f"    pytest {out_dir.name} --html=reports/report.html",
        "",
    ]
    if placeholders:
        lines += [
            "Excel data columns:",
            *[f"  - {p['column_name']}  (sample: {p.get('sample_value','')})" for p in placeholders],
            "",
        ]
    (out_dir / "README.txt").write_text("\n".join(lines), encoding="utf-8")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python -m generator.build_test <recording_dir> <test_name> [--out <dir>]")
        sys.exit(1)

    rec_dir   = Path(sys.argv[1])
    test_name = sys.argv[2]
    out_dir   = config.TESTS_DIR / test_name

    for i, arg in enumerate(sys.argv):
        if arg == "--out" and i + 1 < len(sys.argv):
            out_dir = Path(sys.argv[i + 1])
            break

    generate_test(rec_dir, out_dir, test_name)
